import hashlib
import os
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.utils import timezone
from openpyxl import load_workbook

from audits.models import AuditEvent
import students.models as students
from audits.utils import record_audit_event

REQUIRED_COLUMNS = ("registration_number", "raw_score")
OPTIONAL_COLUMNS = ("grade", "remarks", "status", "is_absent", "is_withheld", "module_code", "batch_code")
REGISTRATION_NUMBER_PATTERN = re.compile(r"^[A-Z0-9]{8,32}$")
STATUS_ALIASES = {"recorded": "recorded", "present": "recorded", "absent": "absent", "withheld": "withheld"}
HEADER_ALIASES = {
    "registration_number": {
        "registration_number",
        "registration_no",
        "registration",
        "reg_number",
        "reg_no",
        "student_id",
        "index_no",
        "index_number",
    },
    "raw_score": {"raw_score", "score", "marks", "mark", "raw_marks", "final_score", "final_mark"},
    "grade": {"grade", "letter_grade", "final_result", "result", "outcome"},
    "remarks": {"remarks", "remark", "comment", "comments", "notes"},
    "status": {"status", "result_status"},
    "is_absent": {"is_absent", "absent", "absence"},
    "is_withheld": {"is_withheld", "withheld"},
    "module_code": {"module_code", "module", "stream_code", "stream", "subject_code", "subject"},
    "batch_code": {"batch_code", "batch", "cohort"},
}


def result_upload_path(instance, filename):
    safe_name = os.path.basename(filename)
    return f"imports/result_uploads/{instance.exam_id or 'unassigned'}/{safe_name}"


class ResultUpload(models.Model):
    class Status(models.TextChoices):
        PENDING = "pending", "Pending"
        STAGED = "staged", "Staged"
        FAILED = "failed", "Failed"
        PUBLISHED = "published", "Published"

    exam = models.ForeignKey("results.Exam", on_delete=models.CASCADE, related_name="uploads")
    uploaded_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.SET_NULL, related_name="result_uploads", blank=True, null=True)
    published_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.SET_NULL, related_name="published_result_uploads", blank=True, null=True)
    source_filename = models.CharField(max_length=255)
    source_file = models.FileField(upload_to=result_upload_path)
    checksum_sha256 = models.CharField(max_length=64)
    version_number = models.PositiveIntegerField(default=1)
    status = models.CharField(max_length=16, choices=Status.choices, default=Status.PENDING)
    total_rows = models.PositiveIntegerField(default=0)
    staged_rows = models.PositiveIntegerField(default=0)
    imported_rows = models.PositiveIntegerField(default=0)
    rejected_rows = models.PositiveIntegerField(default=0)
    started_at = models.DateTimeField(auto_now_add=True)
    staged_at = models.DateTimeField(blank=True, null=True)
    completed_at = models.DateTimeField(blank=True, null=True)
    published_at = models.DateTimeField(blank=True, null=True)
    processing_error = models.TextField(blank=True)
    notes = models.TextField(blank=True)

    class Meta:
        ordering = ["-started_at"]
        constraints = [
            models.UniqueConstraint(fields=["exam", "checksum_sha256"], name="imports_upload_checksum_unique_per_exam"),
            models.UniqueConstraint(fields=["exam", "version_number"], name="imports_upload_version_unique_per_exam"),
            models.CheckConstraint(condition=(models.Q(staged_rows__gte=0) & models.Q(imported_rows__gte=0) & models.Q(rejected_rows__gte=0)), name="imports_upload_non_negative_row_counts"),
        ]

    def clean(self):
        super().clean()
        if self.imported_rows > self.staged_rows:
            raise ValidationError("Imported rows cannot exceed staged rows.")
        if self.staged_rows + self.rejected_rows > self.total_rows:
            raise ValidationError("Staged and rejected row counts cannot exceed total rows.")
        if self.source_file and not self.source_file.name.lower().endswith(".xlsx"):
            raise ValidationError({"source_file": "Only .xlsx workbooks are supported for result imports."})

    def refresh_file_metadata(self):
        if not self.source_file:
            return
        self.source_filename = os.path.basename(self.source_file.name)
        digest = hashlib.sha256()
        self.source_file.open("rb")
        for chunk in self.source_file.chunks():
            digest.update(chunk)
        self.source_file.seek(0)
        self.checksum_sha256 = digest.hexdigest()

    @property
    def ready_rows(self):
        return self.staged_rows - self.imported_rows

    @property
    def summary(self):
        staged_rows = self.staged_rows_set.all()
        return {
            "accepted_rows": staged_rows.filter(review_bucket="accepted").count(),
            "rejected_rows": staged_rows.filter(review_bucket="rejected").count(),
            "duplicate_rows": staged_rows.filter(review_bucket="duplicate").count(),
            "published_rows": self.exam_results.count(),
        }

    def save(self, *args, **kwargs):
        if self._state.adding and self.exam_id:
            latest_version = type(self).objects.filter(exam_id=self.exam_id).aggregate(models.Max("version_number"))["version_number__max"] or 0
            if not self.version_number or self.version_number <= latest_version:
                self.version_number = latest_version + 1
        return super().save(*args, **kwargs)

    def __str__(self):
        return f"v{self.version_number} {self.source_filename} ({self.exam})"


class ResultUploadRow(models.Model):
    class Status(models.TextChoices):
        RECORDED = "recorded", "Recorded"
        ABSENT = "absent", "Absent"
        WITHHELD = "withheld", "Withheld"

    upload = models.ForeignKey(ResultUpload, on_delete=models.CASCADE, related_name="staged_rows_set")
    row_number = models.PositiveIntegerField()
    registration_number = models.CharField(max_length=32, blank=True)
    raw_score = models.DecimalField(max_digits=6, decimal_places=2, blank=True, null=True)
    grade = models.CharField(max_length=16, blank=True)
    remarks = models.TextField(blank=True)
    status = models.CharField(max_length=16, choices=Status.choices, default=Status.RECORDED)
    is_absent = models.BooleanField(default=False)
    student = models.ForeignKey("students.Student", on_delete=models.SET_NULL, related_name="staged_result_rows", blank=True, null=True)
    is_valid = models.BooleanField(default=False)
    review_bucket = models.CharField(max_length=16, blank=True)
    is_duplicate = models.BooleanField(default=False)
    duplicate_reason = models.CharField(max_length=255, blank=True)
    validation_errors = models.JSONField(default=list, blank=True)
    raw_payload = models.JSONField(default=dict, blank=True)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    class Meta:
        ordering = ["row_number"]
        constraints = [models.UniqueConstraint(fields=["upload", "row_number"], name="imports_upload_row_number_unique_per_upload")]

    def __str__(self):
        return f"{self.upload.source_filename} row {self.row_number}"


def _normalize_status(payload):
    raw_status = str(payload.get("status") or "").strip().lower()
    is_absent = str(payload.get("is_absent") or "").strip().lower() in {"1", "true", "yes", "y"}
    is_withheld = str(payload.get("is_withheld") or "").strip().lower() in {"1", "true", "yes", "y"}
    errors = []
    status = STATUS_ALIASES.get(raw_status, "") if raw_status else ""
    if raw_status and not status:
        errors.append("status must be recorded, absent, or withheld.")
    if not status:
        status = ResultUploadRow.Status.ABSENT if is_absent else ResultUploadRow.Status.WITHHELD if is_withheld else ResultUploadRow.Status.RECORDED
    if is_absent and status != ResultUploadRow.Status.ABSENT:
        errors.append("is_absent conflicts with status.")
    if is_withheld and status != ResultUploadRow.Status.WITHHELD:
        errors.append("is_withheld conflicts with status.")
    return status, errors


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _resolve_headers(headers):
    header_index = {}
    unknown_headers = []
    for index, header in enumerate(headers):
        if not header:
            continue
        canonical_name = None
        for candidate, aliases in HEADER_ALIASES.items():
            if header in aliases:
                canonical_name = candidate
                break
        if canonical_name is None:
            unknown_headers.append(header)
            continue
        header_index.setdefault(canonical_name, index)
    return header_index, unknown_headers


def stage_result_upload(upload, *, request=None, channel=AuditEvent.Channel.SERVICE, ip_address=None):
    if upload.exam.is_released:
        message = "Released exams must use the correction workflow instead of import staging."
        record_audit_event(action="imports.results.stage", actor=upload.uploaded_by, instance=upload, request=request, channel=channel, ip_address=ip_address, outcome=AuditEvent.Outcome.FAILED, metadata={"error": message})
        raise ValidationError(message)

    upload.staged_rows_set.all().delete()
    upload.processing_error = ""
    upload.staged_rows = 0
    upload.imported_rows = 0
    upload.rejected_rows = 0
    upload.total_rows = 0
    upload.completed_at = None
    upload.published_at = None
    upload.published_by = None
    upload.refresh_file_metadata()
    upload.full_clean()
    upload.source_file.open("rb")
    workbook_bytes = upload.source_file.read()
    upload.source_file.seek(0)
    workbook = load_workbook(filename=BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        upload.status = ResultUpload.Status.FAILED
        upload.processing_error = "The workbook is empty."
        upload.completed_at = timezone.now()
        upload.save()
        record_audit_event(action="imports.results.stage", actor=upload.uploaded_by, instance=upload, request=request, channel=channel, ip_address=ip_address, outcome=AuditEvent.Outcome.FAILED, metadata={"error": upload.processing_error})
        return upload

    headers = [_normalize_header(value) if value is not None else "" for value in rows[0]]
    header_index, unknown_headers = _resolve_headers(headers)
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in header_index]
    if missing_columns:
        upload.status = ResultUpload.Status.FAILED
        upload.processing_error = "Workbook is missing required columns: " + ", ".join(sorted(missing_columns))
        upload.completed_at = timezone.now()
        upload.save()
        record_audit_event(action="imports.results.stage", actor=upload.uploaded_by, instance=upload, request=request, channel=channel, ip_address=ip_address, outcome=AuditEvent.Outcome.FAILED, metadata={"error": upload.processing_error})
        return upload

    parsed_rows = []
    for row_number, values in enumerate(rows[1:], start=2):
        payload = {}
        for column in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
            idx = header_index.get(column)
            payload[column] = values[idx] if idx is not None and idx < len(values) else None
        registration_number = str(payload.get("registration_number") or "").strip().upper()
        parsed_rows.append((row_number, payload, registration_number))
    upload.total_rows = len(parsed_rows)

    from results.models import ExamResult
    from students.models import Student

    known_students = {
        student.registration_number: student
        for student in Student.objects.select_related("batch").filter(registration_number__in=[registration for _, _, registration in parsed_rows if registration])
    }
    staged_count = 0
    rejected_count = 0
    seen_registration_numbers = set()
    with transaction.atomic():
        upload.save()
        for row_number, payload, registration_number in parsed_rows:
            errors = []
            duplicate_reason = ""
            is_duplicate = False
            raw_score = None
            review_bucket = "accepted"
            status, status_errors = _normalize_status(payload)
            errors.extend(status_errors)

            if all(str(value or "").strip() == "" for value in payload.values()):
                errors.append("Empty row.")
            if not registration_number:
                errors.append("Missing registration_number.")
            elif not REGISTRATION_NUMBER_PATTERN.match(registration_number):
                errors.append("Malformed registration_number.")

            student = known_students.get(registration_number)
            if registration_number and student is None:
                errors.append("Unknown registration_number.")
            if student and student.batch_id != upload.exam.batch_id:
                errors.append("Student batch does not match the exam batch.")

            module_code = str(payload.get("module_code") or "").strip().upper()
            if module_code and module_code != upload.exam.module.code.upper():
                errors.append("Workbook stream or subject code does not match the selected exam stream or subject.")

            batch_code = str(payload.get("batch_code") or "").strip().upper()
            if batch_code and batch_code != upload.exam.batch.code.upper():
                errors.append("Workbook batch_code does not match the selected exam batch.")

            if registration_number in seen_registration_numbers:
                is_duplicate = True
                duplicate_reason = "Duplicate registration_number found in workbook."
                errors.append(duplicate_reason)
            elif registration_number:
                seen_registration_numbers.add(registration_number)

            if student and ExamResult.objects.filter(exam=upload.exam, student=student).exists():
                is_duplicate = True
                duplicate_reason = "An exam result already exists for this student and exam."
                errors.append(duplicate_reason)

            raw_score_value = payload.get("raw_score")
            if status == ResultUploadRow.Status.ABSENT:
                raw_score = None
            elif raw_score_value in (None, ""):
                errors.append("Missing raw_score.")
            else:
                try:
                    raw_score = Decimal(str(raw_score_value)).quantize(Decimal("0.01"))
                except (InvalidOperation, TypeError):
                    errors.append("raw_score must be numeric.")
                else:
                    if raw_score < 0:
                        errors.append("raw_score cannot be negative.")
                    if raw_score > upload.exam.maximum_score:
                        errors.append("raw_score cannot exceed the exam maximum score.")

            if is_duplicate:
                review_bucket = "duplicate"
            elif errors:
                review_bucket = "rejected"

            row = ResultUploadRow.objects.create(
                upload=upload,
                row_number=row_number,
                registration_number=registration_number,
                raw_score=raw_score,
                grade=str(payload.get("grade") or "").strip(),
                remarks=str(payload.get("remarks") or "").strip(),
                status=status,
                is_absent=status == ResultUploadRow.Status.ABSENT,
                student=student,
                is_valid=not errors,
                review_bucket=review_bucket,
                is_duplicate=is_duplicate,
                duplicate_reason=duplicate_reason,
                validation_errors=errors,
                raw_payload={key: "" if value is None else str(value) for key, value in payload.items()},
            )
            if row.is_valid:
                staged_count += 1
            else:
                rejected_count += 1

        upload.staged_rows = staged_count
        upload.rejected_rows = rejected_count
        upload.status = ResultUpload.Status.STAGED if parsed_rows else ResultUpload.Status.FAILED
        upload.staged_at = timezone.now()
        upload.completed_at = timezone.now()
        upload.save()

    record_audit_event(
        action="imports.results.stage",
        actor=upload.uploaded_by,
        instance=upload,
        request=request,
        channel=channel,
        ip_address=ip_address,
        metadata={
            "total_rows": upload.total_rows,
            "staged_rows": upload.staged_rows,
            "rejected_rows": upload.rejected_rows,
            "duplicate_rows": upload.staged_rows_set.filter(review_bucket="duplicate").count(),
            "unknown_headers": unknown_headers,
        },
    )
    return upload


def publish_result_upload(upload, actor=None, *, request=None, channel=AuditEvent.Channel.SERVICE, ip_address=None):
    if upload.exam.is_released:
        message = "Released exams must use the correction workflow instead of re-publishing imports."
        record_audit_event(action="imports.results.publish", actor=actor, instance=upload, request=request, channel=channel, ip_address=ip_address, outcome=AuditEvent.Outcome.FAILED, metadata={"error": message})
        raise ValidationError(message)

    from results.models import ExamResult

    publishable_rows = list(upload.staged_rows_set.select_related("student").filter(is_valid=True, published_result__isnull=True))
    published_count = 0
    with transaction.atomic():
        for row in publishable_rows:
            if ExamResult.objects.filter(exam=upload.exam, student=row.student).exists():
                row.is_valid = False
                row.review_bucket = "duplicate"
                row.is_duplicate = True
                row.duplicate_reason = "An exam result was created after staging and before publication."
                row.validation_errors = list(row.validation_errors) + [row.duplicate_reason]
                row.save(update_fields=["is_valid", "review_bucket", "is_duplicate", "duplicate_reason", "validation_errors", "updated_at"])
                upload.rejected_rows += 1
                continue
            ExamResult.objects.create(
                exam=upload.exam,
                student=row.student,
                upload=upload,
                source_row=row,
                raw_score=row.raw_score,
                grade=row.grade,
                status=row.status,
                is_absent=row.status == ResultUploadRow.Status.ABSENT,
                is_withheld=row.status == ResultUploadRow.Status.WITHHELD,
                remarks=row.remarks,
            )
            published_count += 1
        upload.imported_rows += published_count
        if published_count:
            upload.status = ResultUpload.Status.PUBLISHED
            upload.published_at = timezone.now()
            upload.published_by = actor
        upload.completed_at = timezone.now()
        upload.save()
    record_audit_event(action="imports.results.publish", actor=actor, instance=upload, request=request, channel=channel, ip_address=ip_address, metadata={"published_rows": published_count, "rejected_rows": upload.rejected_rows})
    return published_count



def student_upload_path(instance, filename):
    safe_name = os.path.basename(filename)
    return f"imports/student_uploads/{safe_name}"


class StudentUpload(models.Model):
    class Status(models.TextChoices):
        PENDING = "pending", "Pending"
        STAGED = "staged", "Staged"
        FAILED = "failed", "Failed"
        PUBLISHED = "published", "Published"

    batch = models.ForeignKey("students.Batch", on_delete=models.PROTECT, related_name="student_uploads", blank=True, null=True)
    uploaded_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.SET_NULL, related_name="student_uploads", blank=True, null=True)
    published_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.SET_NULL, related_name="published_student_uploads", blank=True, null=True)
    source_filename = models.CharField(max_length=255)
    source_file = models.FileField(upload_to=student_upload_path)
    checksum_sha256 = models.CharField(max_length=64)
    status = models.CharField(max_length=16, choices=Status.choices, default=Status.PENDING)
    total_rows = models.PositiveIntegerField(default=0)
    staged_rows = models.PositiveIntegerField(default=0)
    imported_rows = models.PositiveIntegerField(default=0)
    rejected_rows = models.PositiveIntegerField(default=0)
    started_at = models.DateTimeField(auto_now_add=True)
    staged_at = models.DateTimeField(blank=True, null=True)
    completed_at = models.DateTimeField(blank=True, null=True)
    published_at = models.DateTimeField(blank=True, null=True)
    processing_error = models.TextField(blank=True)
    notes = models.TextField(blank=True)

    class Meta:
        ordering = ["-started_at"]

    def clean(self):
        super().clean()
        if self.imported_rows > self.staged_rows:
            raise ValidationError("Imported rows cannot exceed staged rows.")
        if self.staged_rows + self.rejected_rows > self.total_rows:
            raise ValidationError("Staged and rejected row counts cannot exceed total rows.")
        if self.source_file and not self.source_file.name.lower().endswith(".xlsx"):
            raise ValidationError({"source_file": "Only .xlsx workbooks are supported for student imports."})

    def refresh_file_metadata(self):
        if not self.source_file:
            return
        self.source_filename = os.path.basename(self.source_file.name)
        digest = hashlib.sha256()
        self.source_file.open("rb")
        for chunk in self.source_file.chunks():
            digest.update(chunk)
        self.source_file.seek(0)
        self.checksum_sha256 = digest.hexdigest()

    @property
    def summary(self):
        staged_rows = self.staged_rows_set.all()
        return {
            "accepted_rows": staged_rows.filter(review_bucket="accepted").count(),
            "rejected_rows": staged_rows.filter(review_bucket="rejected").count(),
            "duplicate_rows": staged_rows.filter(review_bucket="duplicate").count(),
            "published_rows": staged_rows.exclude(created_student__isnull=True).count(),
        }

    def __str__(self):
        return self.source_filename


class StudentUploadRow(models.Model):
    upload = models.ForeignKey(StudentUpload, on_delete=models.CASCADE, related_name="staged_rows_set")
    row_number = models.PositiveIntegerField()
    registration_number = models.CharField(max_length=32, blank=True)
    first_name = models.CharField(max_length=64, blank=True)
    last_name = models.CharField(max_length=64, blank=True)
    university_email = models.EmailField(blank=True)
    status = models.CharField(max_length=16, choices=students.Student.Status.choices, default=students.Student.Status.ACTIVE)
    batch = models.ForeignKey("students.Batch", on_delete=models.SET_NULL, related_name="staged_student_rows", blank=True, null=True)
    matched_student = models.ForeignKey("students.Student", on_delete=models.SET_NULL, related_name="staged_student_import_rows", blank=True, null=True)
    created_student = models.OneToOneField("students.Student", on_delete=models.SET_NULL, related_name="source_student_upload_row", blank=True, null=True)
    is_valid = models.BooleanField(default=False)
    review_bucket = models.CharField(max_length=16, blank=True)
    is_duplicate = models.BooleanField(default=False)
    duplicate_reason = models.CharField(max_length=255, blank=True)
    validation_errors = models.JSONField(default=list, blank=True)
    raw_payload = models.JSONField(default=dict, blank=True)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    class Meta:
        ordering = ["row_number"]
        constraints = [models.UniqueConstraint(fields=["upload", "row_number"], name="imports_student_upload_row_number_unique_per_upload")]

    def __str__(self):
        return f"{self.upload.source_filename} row {self.row_number}"


STUDENT_REQUIRED_COLUMNS = ("registration_number", "first_name", "last_name")
STUDENT_OPTIONAL_COLUMNS = ("university_email", "status")
STUDENT_HEADER_ALIASES = {
    "registration_number": HEADER_ALIASES["registration_number"],
    "first_name": {"first_name", "first", "given_name", "given"},
    "last_name": {"last_name", "last", "surname", "family_name"},
    "university_email": {"university_email", "email", "student_email"},
    "status": {"status", "student_status"},
}
STUDENT_STATUS_ALIASES = {
    "active": students.Student.Status.ACTIVE,
    "leave": students.Student.Status.LEAVE,
    "on_leave": students.Student.Status.LEAVE,
    "graduated": students.Student.Status.GRADUATED,
    "withdrawn": students.Student.Status.WITHDRAWN,
}


def _resolve_student_headers(headers):
    header_index = {}
    unknown_headers = []
    for index, header in enumerate(headers):
        if not header:
            continue
        canonical_name = None
        for candidate, aliases in STUDENT_HEADER_ALIASES.items():
            if header in aliases:
                canonical_name = candidate
                break
        if canonical_name is None:
            unknown_headers.append(header)
            continue
        header_index.setdefault(canonical_name, index)
    return header_index, unknown_headers


def _normalize_student_status(payload):
    raw_status = _normalize_header(payload.get("status")) if payload.get("status") is not None else ""
    if not raw_status:
        return students.Student.Status.ACTIVE, []
    normalized = STUDENT_STATUS_ALIASES.get(raw_status)
    if normalized is None:
        return students.Student.Status.ACTIVE, ["status must be active, leave, graduated, or withdrawn."]
    return normalized, []


def stage_student_upload(upload, *, request=None, channel=AuditEvent.Channel.SERVICE, ip_address=None):
    from django.core.validators import validate_email
    from students.models import Student

    upload.staged_rows_set.all().delete()
    upload.processing_error = ""
    upload.staged_rows = 0
    upload.imported_rows = 0
    upload.rejected_rows = 0
    upload.total_rows = 0
    upload.completed_at = None
    upload.published_at = None
    upload.published_by = None
    upload.refresh_file_metadata()
    upload.full_clean()
    upload.source_file.open("rb")
    workbook_bytes = upload.source_file.read()
    upload.source_file.seek(0)
    workbook = load_workbook(filename=BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        upload.status = StudentUpload.Status.FAILED
        upload.processing_error = "The workbook is empty."
        upload.completed_at = timezone.now()
        upload.save()
        record_audit_event(action="imports.students.stage", actor=upload.uploaded_by, instance=upload, request=request, channel=channel, ip_address=ip_address, outcome=AuditEvent.Outcome.FAILED, metadata={"error": upload.processing_error})
        return upload

    headers = [_normalize_header(value) if value is not None else "" for value in rows[0]]
    header_index, unknown_headers = _resolve_student_headers(headers)
    missing_columns = [column for column in STUDENT_REQUIRED_COLUMNS if column not in header_index]
    if missing_columns:
        upload.status = StudentUpload.Status.FAILED
        upload.processing_error = "Workbook is missing required columns: " + ", ".join(sorted(missing_columns))
        upload.completed_at = timezone.now()
        upload.save()
        record_audit_event(action="imports.students.stage", actor=upload.uploaded_by, instance=upload, request=request, channel=channel, ip_address=ip_address, outcome=AuditEvent.Outcome.FAILED, metadata={"error": upload.processing_error})
        return upload

    parsed_rows = []
    for row_number, values in enumerate(rows[1:], start=2):
        payload = {}
        for column in STUDENT_REQUIRED_COLUMNS + STUDENT_OPTIONAL_COLUMNS:
            idx = header_index.get(column)
            payload[column] = values[idx] if idx is not None and idx < len(values) else None
        registration_number = str(payload.get("registration_number") or "").strip().upper()
        parsed_rows.append((row_number, payload, registration_number))
    upload.total_rows = len(parsed_rows)

    existing_students = {student.registration_number: student for student in Student.objects.filter(registration_number__in=[registration for _, _, registration in parsed_rows if registration])}
    staged_count = 0
    rejected_count = 0
    seen_registration_numbers = set()
    with transaction.atomic():
        upload.save()
        for row_number, payload, registration_number in parsed_rows:
            errors = []
            duplicate_reason = ""
            is_duplicate = False
            review_bucket = "accepted"
            first_name = str(payload.get("first_name") or "").strip()
            last_name = str(payload.get("last_name") or "").strip()
            university_email = str(payload.get("university_email") or "").strip().lower()
            status, status_errors = _normalize_student_status(payload)
            errors.extend(status_errors)

            if all(str(value or "").strip() == "" for value in payload.values()):
                errors.append("Empty row.")
            if not registration_number:
                errors.append("Missing registration_number.")
            elif not REGISTRATION_NUMBER_PATTERN.match(registration_number):
                errors.append("Malformed registration_number.")
            if not first_name:
                errors.append("Missing first_name.")
            if not last_name:
                errors.append("Missing last_name.")
            if university_email:
                try:
                    validate_email(university_email)
                except ValidationError:
                    errors.append("university_email must be valid.")

            matched_student = existing_students.get(registration_number)
            if registration_number in seen_registration_numbers:
                is_duplicate = True
                duplicate_reason = "Duplicate registration_number found in workbook."
                errors.append(duplicate_reason)
            elif registration_number:
                seen_registration_numbers.add(registration_number)
            if matched_student is not None:
                is_duplicate = True
                duplicate_reason = "A student with this registration_number already exists."
                errors.append(duplicate_reason)

            if is_duplicate:
                review_bucket = "duplicate"
            elif errors:
                review_bucket = "rejected"

            row = StudentUploadRow.objects.create(
                upload=upload,
                row_number=row_number,
                registration_number=registration_number,
                first_name=first_name,
                last_name=last_name,
                university_email=university_email,
                status=status,
                batch=upload.batch,
                matched_student=matched_student,
                is_valid=not errors,
                review_bucket=review_bucket,
                is_duplicate=is_duplicate,
                duplicate_reason=duplicate_reason,
                validation_errors=errors,
                raw_payload={key: "" if value is None else str(value) for key, value in payload.items()},
            )
            if row.is_valid:
                staged_count += 1
            else:
                rejected_count += 1

        upload.staged_rows = staged_count
        upload.rejected_rows = rejected_count
        upload.status = StudentUpload.Status.STAGED if parsed_rows else StudentUpload.Status.FAILED
        upload.staged_at = timezone.now()
        upload.completed_at = timezone.now()
        upload.save()

    record_audit_event(action="imports.students.stage", actor=upload.uploaded_by, instance=upload, request=request, channel=channel, ip_address=ip_address, metadata={"total_rows": upload.total_rows, "staged_rows": upload.staged_rows, "rejected_rows": upload.rejected_rows, "duplicate_rows": upload.staged_rows_set.filter(review_bucket="duplicate").count(), "unknown_headers": unknown_headers, "batch_id": upload.batch_id})
    return upload


def publish_student_upload(upload, actor=None, *, request=None, channel=AuditEvent.Channel.SERVICE, ip_address=None):
    from students.models import Student

    publishable_rows = list(upload.staged_rows_set.select_related("batch").filter(is_valid=True, created_student__isnull=True))
    published_count = 0
    with transaction.atomic():
        for row in publishable_rows:
            if Student.objects.filter(registration_number=row.registration_number).exists():
                row.is_valid = False
                row.review_bucket = "duplicate"
                row.is_duplicate = True
                row.duplicate_reason = "A student was created after staging and before publication."
                row.validation_errors = list(row.validation_errors) + [row.duplicate_reason]
                row.save(update_fields=["is_valid", "review_bucket", "is_duplicate", "duplicate_reason", "validation_errors", "updated_at"])
                upload.rejected_rows += 1
                continue
            student = Student.objects.create(
                batch=row.batch,
                registration_number=row.registration_number,
                first_name=row.first_name,
                last_name=row.last_name,
                university_email=row.university_email,
                status=row.status,
            )
            row.created_student = student
            row.save(update_fields=["created_student", "updated_at"])
            published_count += 1
        upload.imported_rows += published_count
        if published_count:
            upload.status = StudentUpload.Status.PUBLISHED
            upload.published_at = timezone.now()
            upload.published_by = actor
        upload.completed_at = timezone.now()
        upload.save()
    record_audit_event(action="imports.students.publish", actor=actor, instance=upload, request=request, channel=channel, ip_address=ip_address, metadata={"published_rows": published_count, "rejected_rows": upload.rejected_rows, "batch_id": upload.batch_id})
    return published_count

