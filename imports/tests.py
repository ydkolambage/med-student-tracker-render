import hashlib
import shutil
import tempfile
from decimal import Decimal
from io import BytesIO

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from imports.admin import ResultUploadAdmin
from imports.models import ResultUpload, StudentUpload, publish_result_upload, stage_result_upload, stage_student_upload
from results.models import Exam, ExamResult
from results.services import release_exam_results
from students.models import Batch, Department, Module, Student


class ImportPipelineTests(TestCase):
    def setUp(self):
        self.temp_media = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.temp_media)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.addCleanup(lambda: shutil.rmtree(self.temp_media, ignore_errors=True))
        self.user = get_user_model().objects.create_user(username="importer", email="importer@example.com", password="safe-password-123")
        self.department = Department.objects.create(code="MED", name="Medicine")
        self.batch = Batch.objects.create(code="MBBS-2025", display_name="MBBS 2025 Cohort", academic_start_year=2025)
        self.module = Module.objects.create(department=self.department, code="ANA101", title="Anatomy I", semester="1")
        self.module.batches.set([self.batch])
        self.exam = Exam.objects.create(batch=self.batch, module=self.module, title="Final anatomy exam", sat_on="2026-05-11", maximum_score=Decimal("100.00"), pass_mark=Decimal("50.00"))
        self.student = Student.objects.create(batch=self.batch, registration_number="2025MED0001", first_name="Asha", last_name="Silva")
        self.other_student = Student.objects.create(batch=self.batch, registration_number="2025MED0002", first_name="Bimal", last_name="Fernando")
        self.other_batch = Batch.objects.create(code="MBBS-2024", display_name="MBBS 2024 Cohort", academic_start_year=2024)
        self.other_batch_student = Student.objects.create(batch=self.other_batch, registration_number="2024MED0001", first_name="Chanu", last_name="Perera")

    def build_workbook_file(self, rows, filename="results.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(filename, buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def create_upload(self, workbook_rows, filename="results.xlsx"):
        return ResultUpload.objects.create(exam=self.exam, uploaded_by=self.user, source_filename=filename, source_file=self.build_workbook_file(workbook_rows, filename=filename), checksum_sha256="0" * 64)

    def test_stage_upload_rejects_missing_required_columns(self):
        upload = self.create_upload([["registration_number", "grade"], [self.student.registration_number, "A"]])
        stage_result_upload(upload)
        upload.refresh_from_db()
        self.assertEqual(upload.status, ResultUpload.Status.FAILED)
        self.assertIn("raw_score", upload.processing_error)
        self.assertEqual(upload.staged_rows_set.count(), 0)

    def test_stage_upload_detects_duplicates_and_status_conflicts(self):
        ExamResult.objects.create(exam=self.exam, student=self.other_student, raw_score=Decimal("78.00"), grade="B+")
        upload = self.create_upload([["registration_number", "raw_score", "status", "is_absent"], [self.student.registration_number, 74, "recorded", False], [self.student.registration_number, 75, "recorded", False], ["bad-id", 80, "absent", False], [self.other_student.registration_number, 70, "recorded", False]], filename="duplicates.xlsx")
        stage_result_upload(upload)
        upload.refresh_from_db()
        rows = {row.row_number: row for row in upload.staged_rows_set.all()}
        self.assertEqual(upload.status, ResultUpload.Status.STAGED)
        self.assertEqual(upload.staged_rows, 1)
        self.assertEqual(upload.rejected_rows, 3)
        self.assertTrue(rows[3].is_duplicate)
        self.assertIn("Duplicate registration_number", rows[3].validation_errors[0])
        self.assertIn("Malformed registration_number.", rows[4].validation_errors)
        self.assertTrue(rows[5].is_duplicate)

    def test_stage_upload_supports_header_aliases_and_versions(self):
        upload = self.create_upload([["Registration Number", "Marks", "Module", "Batch"], [self.student.registration_number, 88, self.module.code, self.batch.code]], filename="aliases-v1.xlsx")
        stage_result_upload(upload)
        upload.refresh_from_db()
        row = upload.staged_rows_set.get()
        self.assertEqual(upload.version_number, 1)
        self.assertEqual(upload.staged_rows, 1)
        self.assertEqual(row.review_bucket, "accepted")
        self.assertEqual(row.raw_score, Decimal("88.00"))

        second_upload = self.create_upload([["Reg No", "Score"], [self.other_student.registration_number, 66]], filename="aliases-v2.xlsx")
        stage_result_upload(second_upload)
        second_upload.refresh_from_db()
        self.assertEqual(second_upload.version_number, 2)

    def test_stage_upload_accepts_final_score_and_final_result_headers(self):
        upload = self.create_upload([["registration_number", "final_score", "final_result"], [self.student.registration_number, 81, "Pass"]], filename="final-sheet.xlsx")
        stage_result_upload(upload)
        upload.refresh_from_db()
        row = upload.staged_rows_set.get()
        self.assertTrue(row.is_valid)
        self.assertEqual(row.raw_score, Decimal("81.00"))
        self.assertEqual(row.grade, "Pass")

    def test_stage_upload_flags_empty_rows_and_batch_mismatches(self):
        upload = self.create_upload([["registration_number", "raw_score", "module_code", "batch_code"], ["", "", "", ""], [self.other_batch_student.registration_number, 71, self.module.code, self.other_batch.code]], filename="mismatch.xlsx")
        stage_result_upload(upload)
        rows = {row.row_number: row for row in upload.staged_rows_set.all()}
        self.assertIn("Empty row.", rows[2].validation_errors)
        self.assertIn("Student batch does not match the exam batch.", rows[3].validation_errors)
        self.assertIn("Workbook batch_code does not match the selected exam batch.", rows[3].validation_errors)
        self.assertEqual(rows[2].review_bucket, "rejected")

    def test_admin_save_model_stores_uploader_identity_and_file_hash(self):
        admin = ResultUploadAdmin(ResultUpload, AdminSite())
        request = RequestFactory().post("/admin/imports/resultupload/add/")
        request.user = self.user
        workbook = self.build_workbook_file([["registration_number", "raw_score"], [self.student.registration_number, 88]], filename="hash-check.xlsx")
        expected_hash = hashlib.sha256(workbook.read()).hexdigest()
        workbook.seek(0)
        upload = ResultUpload(exam=self.exam, source_filename="hash-check.xlsx", source_file=workbook, checksum_sha256="")
        admin.save_model(request, upload, form=None, change=False)
        upload.refresh_from_db()
        self.assertEqual(upload.uploaded_by, self.user)
        self.assertEqual(upload.checksum_sha256, expected_hash)
        self.assertEqual(upload.source_filename, upload.source_file.name.split("/")[-1])

    def test_results_are_not_published_automatically(self):
        upload = self.create_upload([["registration_number", "raw_score", "grade"], [self.student.registration_number, 91, "A"]], filename="publish.xlsx")
        stage_result_upload(upload)
        upload.refresh_from_db()
        self.assertEqual(ExamResult.objects.count(), 0)
        self.assertEqual(upload.status, ResultUpload.Status.STAGED)
        published_count = publish_result_upload(upload, actor=self.user)
        upload.refresh_from_db()
        staged_row = upload.staged_rows_set.get()
        result = ExamResult.objects.get()
        self.assertEqual(published_count, 1)
        self.assertEqual(upload.status, ResultUpload.Status.PUBLISHED)
        self.assertEqual(upload.imported_rows, 1)
        self.assertEqual(upload.published_by, self.user)
        self.assertEqual(result.source_row, staged_row)
        self.assertEqual(result.raw_score, Decimal("91.00"))

    def test_stage_and_publish_are_blocked_after_release(self):
        upload = self.create_upload([["registration_number", "raw_score", "grade"], [self.student.registration_number, 91, "A"]], filename="released.xlsx")
        ExamResult.objects.create(exam=self.exam, student=self.student, raw_score=Decimal("65.00"), grade="B")
        release_exam_results(self.exam, actor=self.user)
        with self.assertRaises(ValidationError):
            stage_result_upload(upload)
        with self.assertRaises(ValidationError):
            publish_result_upload(upload, actor=self.user)


class ResultImportWorkflowTests(TestCase):
    def setUp(self):
        self.temp_media = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.temp_media)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.addCleanup(lambda: shutil.rmtree(self.temp_media, ignore_errors=True))
        self.user = get_user_model().objects.create_superuser(username="resultadmin", email="resultadmin@example.com", password="safe-password-123")
        self.department = Department.objects.create(code="BSC", name="Basic Sciences")
        self.batch = Batch.objects.create(code="BATCH-1", display_name="Batch 1", academic_start_year=2025)
        self.module = Module.objects.create(department=self.department, code="2NDMBBS", title="2nd MBBS", semester="1-3")
        self.module.batches.set([self.batch])
        self.exam = Exam.objects.create(batch=self.batch, module=self.module, title="2nd MBBS Examination", sat_on="2026-04-15")
        self.student = Student.objects.create(batch=self.batch, registration_number="BSC20250001", first_name="Aarav", last_name="Perera")
        self.client.force_login(self.user)

    def build_workbook(self, rows, filename="results.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(filename, buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_imports_home_lists_student_and_result_sections(self):
        response = self.client.get(reverse("imports-home"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Student Import")
        self.assertContains(response, "Results Import")

    def test_result_import_form_stages_upload_using_ui_selected_batch_stream_and_exam(self):
        workbook = self.build_workbook([["registration_number", "final_score", "final_result"], [self.student.registration_number, 76, "Pass"]], filename="2nd-mbbs.xlsx")
        response = self.client.post(reverse("result-import-new"), {"batch_id": str(self.batch.id), "module_id": str(self.module.id), "exam_id": str(self.exam.id), "source_file": workbook, "notes": "semester finals"})
        self.assertEqual(response.status_code, 302)
        upload = ResultUpload.objects.get()
        self.assertEqual(upload.exam, self.exam)
        row = upload.staged_rows_set.get()
        self.assertEqual(row.raw_score, Decimal("76.00"))
        self.assertEqual(row.grade, "Pass")

    def test_result_import_form_requires_exam_selection_or_new_exam_name(self):
        workbook = self.build_workbook([["registration_number", "final_score", "final_result"], [self.student.registration_number, 76, "Pass"]], filename="2nd-mbbs.xlsx")
        response = self.client.post(reverse("result-import-new"), {"batch_id": str(self.batch.id), "module_id": str(self.module.id), "source_file": workbook, "notes": "semester finals"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Choose an exam or enter a new exam name before uploading.")
        self.assertEqual(ResultUpload.objects.count(), 0)

    def test_result_import_form_can_create_exam_when_not_listed(self):
        workbook = self.build_workbook([["registration_number", "final_score", "final_result"], [self.student.registration_number, 76, "Pass"]], filename="2nd-mbbs.xlsx")
        response = self.client.post(reverse("result-import-new"), {"batch_id": str(self.batch.id), "module_id": str(self.module.id), "new_exam_title": "Repeat assessment", "new_exam_sat_on": "2026-05-01", "source_file": workbook, "notes": "semester finals"})
        self.assertEqual(response.status_code, 302)
        created_exam = Exam.objects.get(title="Repeat assessment")
        self.assertEqual(created_exam.batch, self.batch)
        self.assertEqual(created_exam.module, self.module)
        self.assertEqual(created_exam.sat_on.isoformat(), "2026-05-01")
        self.assertEqual(created_exam.maximum_score, Decimal("100.00"))
        self.assertEqual(created_exam.pass_mark, Decimal("50.00"))
        upload = ResultUpload.objects.get()
        self.assertEqual(upload.exam, created_exam)


class StudentImportWorkflowTests(TestCase):
    def setUp(self):
        self.temp_media = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.temp_media)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.addCleanup(lambda: shutil.rmtree(self.temp_media, ignore_errors=True))
        self.user = get_user_model().objects.create_superuser(username="admin", email="admin@example.com", password="safe-password-123")
        self.batch = Batch.objects.create(code="BATCH-1", display_name="Batch 1", academic_start_year=2025)
        self.other_batch = Batch.objects.create(code="BATCH-2", display_name="Batch 2", academic_start_year=2024)
        self.existing_student = Student.objects.create(batch=self.batch, registration_number="BSC20250001", first_name="Existing", last_name="Student")
        self.client.force_login(self.user)

    def build_workbook(self, rows, filename="students.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(filename, buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_stage_student_upload_detects_duplicates_and_missing_names(self):
        upload = StudentUpload.objects.create(batch=self.batch, uploaded_by=self.user, source_filename="students.xlsx", source_file=self.build_workbook([["registration_number", "first_name", "last_name"], ["BSC20250002", "Asha", "Silva"], ["BSC20250002", "Asha", "Silva"], [self.existing_student.registration_number, "Existing", "Student"], ["BSC20250003", "", "Fernando"]]), checksum_sha256="0" * 64)
        stage_student_upload(upload)
        upload.refresh_from_db()
        rows = {row.row_number: row for row in upload.staged_rows_set.all()}
        self.assertEqual(upload.staged_rows, 1)
        self.assertEqual(upload.rejected_rows, 3)
        self.assertTrue(rows[3].is_duplicate)
        self.assertTrue(rows[4].is_duplicate)
        self.assertIn("Missing first_name.", rows[5].validation_errors)
        self.assertEqual(rows[2].batch, self.batch)

    def test_stage_student_upload_accepts_optional_email_and_ui_selected_batch(self):
        upload = StudentUpload.objects.create(batch=self.other_batch, uploaded_by=self.user, source_filename="students.xlsx", source_file=self.build_workbook([["registration_number", "first_name", "last_name", "university_email"], ["BSC20250011", "Nila", "Perera", ""]]), checksum_sha256="0" * 64)
        stage_student_upload(upload)
        row = upload.staged_rows_set.get()
        self.assertTrue(row.is_valid)
        self.assertEqual(row.batch, self.other_batch)
        self.assertEqual(row.university_email, "")

    def test_publish_summary_creates_students(self):
        workbook = self.build_workbook([["registration_number", "first_name", "last_name", "status", "university_email"], ["BSC20250010", "New", "Student", "active", "student10@example.edu"]], filename="students.xlsx")
        upload = StudentUpload.objects.create(batch=self.batch, uploaded_by=self.user, source_filename="students.xlsx", source_file=workbook, checksum_sha256="0" * 64)
        stage_student_upload(upload)
        response = self.client.post(reverse("student-upload-publish-summary", args=[upload.id]))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Student.objects.filter(registration_number="BSC20250010", batch=self.batch).exists())
        upload.refresh_from_db()
        self.assertEqual(upload.status, StudentUpload.Status.PUBLISHED)

    def test_student_import_form_requires_selected_batch(self):
        workbook = self.build_workbook([["registration_number", "first_name", "last_name"], ["BSC20250021", "Form", "Only"]], filename="students.xlsx")
        response = self.client.post(reverse("student-import-new"), {"source_file": workbook, "notes": ""})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Choose a batch before uploading.")
        self.assertEqual(StudentUpload.objects.count(), 0)

    def test_student_import_form_shows_clean_batch_numbers(self):
        response = self.client.get(reverse("student-import-new"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'<option value="{self.batch.id}" >1</option>', html=True)
        self.assertNotContains(response, "? Batch")

    def test_student_import_form_can_create_new_batch_number(self):
        workbook = self.build_workbook([["registration_number", "first_name", "last_name"], ["BSC20250030", "Fresh", "Student"]], filename="students.xlsx")
        response = self.client.post(reverse("student-import-new"), {"source_file": workbook, "notes": "", "new_batch_number": "3"})
        self.assertEqual(response.status_code, 302)
        created_batch = Batch.objects.get(code="BATCH-3")
        self.assertEqual(created_batch.display_name, "Batch 3")
        upload = StudentUpload.objects.get()
        self.assertEqual(upload.batch, created_batch)

    def test_student_review_export_honors_duplicate_bucket(self):
        workbook = self.build_workbook([["registration_number", "first_name", "last_name"], ["BSC20250012", "One", "Student"], ["BSC20250012", "Two", "Student"]], filename="students.xlsx")
        upload = StudentUpload.objects.create(batch=self.batch, uploaded_by=self.user, source_filename="students.xlsx", source_file=workbook, checksum_sha256="0" * 64)
        stage_student_upload(upload)
        response = self.client.get(reverse("student-upload-review-export", args=[upload.id]), {"bucket": "duplicate"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("Duplicate registration_number found in workbook.", response.content.decode())

